from __future__ import annotations

import json
import os
import sqlite3
import tempfile
import uuid
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


class AuctionStorageError(RuntimeError):
    pass


class AuctionStorage:
    EXCEL_COLUMNS = (
        "Auction Date", "Exit Market/Storage", "Entry Market/Storage",
        "Capacity Type", "Network Point Name", "Product Type", "Flow Start",
        "Flow End", "Booked Capacity, kWh/h", "Runtime Hours",
        "Tariff, EUR/MWh/h", "Premium, EUR/MWh/h", "Auction ID",
        "TSO Exit", "TSO Entry", "Status",
    )
    def __init__(self, database_path: Path) -> None:
        database_path.parent.mkdir(parents=True, exist_ok=True)
        self.database_path = database_path
        self._create_schema()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.database_path)
        connection.row_factory = sqlite3.Row
        return connection

    def _create_schema(self) -> None:
        with self._connect() as connection:
            connection.executescript("""
                CREATE TABLE IF NOT EXISTS auctions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    auction_id TEXT NOT NULL, auction_date TEXT NOT NULL,
                    exit_market TEXT NOT NULL DEFAULT '', entry_market TEXT NOT NULL DEFAULT '',
                    direction TEXT NOT NULL, network_point TEXT NOT NULL,
                    network_point_id TEXT NOT NULL DEFAULT '', tso_exit TEXT NOT NULL DEFAULT '',
                    tso_entry TEXT NOT NULL DEFAULT '', product_type TEXT NOT NULL,
                    flow_start TEXT NOT NULL, flow_end TEXT NOT NULL,
                    booked_capacity_kwh_h REAL NOT NULL, runtime_hours REAL NOT NULL,
                    tariff_eur_mwh_h REAL NOT NULL, premium_eur_mwh_h REAL NOT NULL,
                    state TEXT NOT NULL DEFAULT '', created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE (auction_id, network_point_id, direction, flow_start, flow_end)
                );
                CREATE TABLE IF NOT EXISTS prisma_source_operations (
                    operation_id TEXT PRIMARY KEY,
                    source_date TEXT NOT NULL,
                    source_name TEXT NOT NULL,
                    sha256 TEXT NOT NULL,
                    status TEXT NOT NULL CHECK(status IN ('pending','data_committed','accepted')),
                    summary_json TEXT,
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE(source_date)
                );
            """)

    def operations(self) -> list[sqlite3.Row]:
        with self._connect() as connection:
            return list(connection.execute(
                "SELECT * FROM prisma_source_operations ORDER BY source_date"
            ))

    def unresolved_operations(self) -> list[sqlite3.Row]:
        return [row for row in self.operations() if row["status"] != "accepted"]

    def import_legacy_operation(
        self, operation_id: str, source_date: str, source_name: str, digest: str
    ) -> None:
        with self._connect() as connection:
            connection.execute(
                "INSERT OR IGNORE INTO prisma_source_operations "
                "(operation_id, source_date, source_name, sha256, status) "
                "VALUES (?, ?, ?, ?, 'accepted')",
                (operation_id, source_date, source_name, digest),
            )

    def operation_for_date(self, source_date: str) -> sqlite3.Row | None:
        with self._connect() as connection:
            return connection.execute(
                "SELECT * FROM prisma_source_operations WHERE source_date = ?", (source_date,)
            ).fetchone()

    def begin_operation(self, source_date: str, source_name: str, digest: str) -> sqlite3.Row:
        existing = self.operation_for_date(source_date)
        if existing is not None:
            if existing["sha256"] != digest:
                state = "accepted" if existing["status"] == "accepted" else "unresolved"
                raise AuctionStorageError(
                    f"A different PRISMA source for this date is already {state}."
                )
            return existing
        operation_id = uuid.uuid4().hex
        with self._connect() as connection:
            connection.execute(
                "INSERT INTO prisma_source_operations "
                "(operation_id, source_date, source_name, sha256, status) VALUES (?, ?, ?, ?, 'pending')",
                (operation_id, source_date, source_name, digest),
            )
        return self.operation_for_date(source_date)  # type: ignore[return-value]

    def apply_operation(self, operation_id: str, rows: list[dict[str, Any]], summary: dict[str, Any]) -> dict[str, int]:
        with self._connect() as connection:
            operation = connection.execute(
                "SELECT status, summary_json FROM prisma_source_operations WHERE operation_id = ?",
                (operation_id,),
            ).fetchone()
            if operation is None:
                raise AuctionStorageError("The pending PRISMA operation was not found.")
            if operation["status"] != "pending":
                stored = json.loads(operation["summary_json"] or "{}")
                return {key: int(stored[key]) for key in ("processed", "inserted", "updated", "unchanged")}
            stats = self._upsert_rows(connection, rows)
            summary.update(stats)
            connection.execute(
                "UPDATE prisma_source_operations SET status='data_committed', summary_json=?, "
                "updated_at=CURRENT_TIMESTAMP WHERE operation_id=? AND status='pending'",
                (json.dumps(summary, sort_keys=True), operation_id),
            )
        return stats

    @staticmethod
    def _upsert_rows(
        connection: sqlite3.Connection, rows: list[dict[str, Any]]
    ) -> dict[str, int]:
        inserted = updated = unchanged = 0
        for row in rows:
            existing = connection.execute(
                "SELECT * FROM auctions WHERE auction_id=? AND network_point_id=? "
                "AND direction=? AND flow_start=? AND flow_end=?",
                (row["auction_id"], row["network_point_id"], row["direction"],
                 row["flow_start"], row["flow_end"]),
            ).fetchone()
            if existing is None:
                columns = ", ".join(row)
                connection.execute(
                    f"INSERT INTO auctions ({columns}) VALUES ({', '.join('?' for _ in row)})",
                    tuple(row.values()),
                )
                inserted += 1
            elif any(
                existing[key] != value for key, value in row.items()
                if key in existing.keys()
            ):
                assignments = ", ".join(f"{key}=?" for key in row)
                connection.execute(
                    f"UPDATE auctions SET {assignments}, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                    (*row.values(), existing["id"]),
                )
                updated += 1
            else:
                unchanged += 1
        return {"processed": len(rows), "inserted": inserted,
                "updated": updated, "unchanged": unchanged}

    def finalize_operation(self, operation_id: str) -> None:
        with self._connect() as connection:
            changed = connection.execute(
                "UPDATE prisma_source_operations SET status='accepted', updated_at=CURRENT_TIMESTAMP "
                "WHERE operation_id=? AND status='data_committed'", (operation_id,)
            ).rowcount
            if changed != 1:
                raise AuctionStorageError("The PRISMA operation could not be finalized safely.")

    @staticmethod
    def validate_excel(path: Path) -> bool:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            valid = False
            if "Auctions" in workbook.sheetnames:
                sheet = workbook["Auctions"]
                headers = tuple(
                    cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))
                )
                valid = headers == AuctionStorage.EXCEL_COLUMNS
            workbook.close()
            return valid
        except Exception:
            return False

    def export_excel(self, output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with self._connect() as connection:
            frame = pd.read_sql_query("""
                SELECT auction_date AS "Auction Date", exit_market AS "Exit Market/Storage",
                entry_market AS "Entry Market/Storage", direction AS "Capacity Type",
                network_point AS "Network Point Name", product_type AS "Product Type",
                flow_start AS "Flow Start", flow_end AS "Flow End",
                booked_capacity_kwh_h AS "Booked Capacity, kWh/h", runtime_hours AS "Runtime Hours",
                tariff_eur_mwh_h AS "Tariff, EUR/MWh/h", premium_eur_mwh_h AS "Premium, EUR/MWh/h",
                auction_id AS "Auction ID", tso_exit AS "TSO Exit", tso_entry AS "TSO Entry", state AS "Status"
                FROM auctions ORDER BY auction_date, auction_id, network_point_id, direction, flow_start, flow_end
            """, connection)
        staged: Path | None = None
        try:
            descriptor, name = tempfile.mkstemp(prefix=f".{output_path.stem}-", suffix=".xlsx", dir=output_path.parent)
            os.close(descriptor)
            staged = Path(name)
            frame.to_excel(staged, index=False, sheet_name="Auctions")
            if not self.validate_excel(staged):
                raise AuctionStorageError("The staged Excel workbook failed validation.")
            try:
                os.replace(staged, output_path)
            except PermissionError as exc:
                raise AuctionStorageError(
                    "The Excel output is open or locked. Close it and retry the import."
                ) from exc
            staged = None
        except AuctionStorageError:
            raise
        except Exception as exc:
            raise AuctionStorageError("The Excel output could not be staged safely.") from exc
        finally:
            if staged is not None:
                try:
                    staged.unlink(missing_ok=True)
                except OSError:
                    pass
        return output_path

    def upsert(self, rows: list[dict[str, Any]]) -> dict[str, int]:
        """Compatibility API for storage-only callers."""
        with self._connect() as connection:
            return self._upsert_rows(connection, rows)
