import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from storage import AuctionStorage


APPROVED_EXCEL_COLUMN_WIDTHS = {
    "Auction Date": 21,
    "Exit Market/Storage": 22,
    "Entry Market/Storage": 22,
    "Capacity Type": 15,
    "Network Point Name": 36,
    "Product Type": 14,
    "Flow Start": 21,
    "Flow End": 21,
    "Booked Capacity, kWh/h": 24,
    "Runtime Hours": 15,
    "Tariff, EUR/MWh/h": 20,
    "Premium, EUR/MWh/h": 21,
    "Auction ID": 16,
    "TSO Exit": 30,
    "TSO Entry": 30,
    "Status": 14,
}


def test_duplicate_import_does_not_insert_twice(tmp_path) -> None:
    storage = AuctionStorage(tmp_path / "test.db")
    row = {
        "auction_id": "123",
        "auction_date": "2026-07-10T06:00:00",
        "exit_market": "",
        "entry_market": "",
        "direction": "entry",
        "network_point": "Test Point",
        "network_point_id": "NP-1",
        "tso_exit": "",
        "tso_entry": "Test TSO",
        "product_type": "Day Ahead",
        "flow_start": "2026-07-10T06:00:00",
        "flow_end": "2026-07-11T06:00:00",
        "booked_capacity_kwh_h": 1000.0,
        "runtime_hours": 24.0,
        "tariff_eur_mwh_h": 1.0,
        "premium_eur_mwh_h": 0.0,
        "state": "Finished",
    }

    first = storage.upsert([row])
    second = storage.upsert([row])

    assert first["inserted"] == 1
    assert second["inserted"] == 0
    assert second["unchanged"] == 1


def test_excel_width_mapping_covers_columns_and_header_only_export(tmp_path) -> None:
    storage = AuctionStorage(tmp_path / "test.db")
    output = storage.export_excel(tmp_path / "result.xlsx")

    assert AuctionStorage.EXCEL_COLUMN_WIDTHS == APPROVED_EXCEL_COLUMN_WIDTHS
    assert tuple(APPROVED_EXCEL_COLUMN_WIDTHS) == AuctionStorage.EXCEL_COLUMNS
    assert AuctionStorage.validate_excel(output)
    workbook = load_workbook(output)
    sheet = workbook["Auctions"]
    assert sheet.max_row == 1
    assert {
        header: sheet.column_dimensions[get_column_letter(index)].width
        for index, header in enumerate(AuctionStorage.EXCEL_COLUMNS, start=1)
    } == APPROVED_EXCEL_COLUMN_WIDTHS
    workbook.close()


def test_populated_excel_export_has_production_widths(tmp_path) -> None:
    storage = AuctionStorage(tmp_path / "test.db")
    storage.upsert([{
        "auction_id": "123", "auction_date": "2026-07-10T06:00:00",
        "exit_market": "", "entry_market": "VGS Storage Hub",
        "direction": "entry", "network_point": "VGS Storage Hub (4290)",
        "network_point_id": "NP-1", "tso_exit": "", "tso_entry": "Test TSO",
        "product_type": "Day Ahead", "flow_start": "2026-07-10T06:00:00",
        "flow_end": "2026-07-11T06:00:00", "booked_capacity_kwh_h": 1000.0,
        "runtime_hours": 24.0, "tariff_eur_mwh_h": 1.0,
        "premium_eur_mwh_h": 0.0, "state": "Finished",
    }])

    output = storage.export_excel(tmp_path / "result.xlsx")

    assert AuctionStorage.validate_excel(output)
    assert pd.read_excel(output)["Auction ID"].astype(str).tolist() == ["123"]


def test_excel_validation_rejects_missing_or_incorrect_widths(tmp_path) -> None:
    path = tmp_path / "result.xlsx"
    pd.DataFrame(columns=AuctionStorage.EXCEL_COLUMNS).to_excel(
        path, index=False, sheet_name="Auctions"
    )
    assert not AuctionStorage.validate_excel(path)

    AuctionStorage.apply_excel_widths(path)
    workbook = load_workbook(path)
    workbook["Auctions"].column_dimensions["A"].width = 12
    workbook.save(path)
    workbook.close()

    assert not AuctionStorage.validate_excel(path)
